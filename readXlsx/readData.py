import openpyxl
from tabulate import tabulate

ruta = "readXlsx\pruebasubirArchivos.xlsx"

def readData(ruta):

     # Define variable to load the dataframe
     excel_dataframe = openpyxl.load_workbook(ruta)

     # Define variable to read sheet
     dataframe = excel_dataframe.active

     print(dataframe)

     encabezado=[]
     data = []




     # se definen las variables de entradas y salidas de la matriz
     entradas=0;
     salidas=0;
     valoresNulos=0;
     # recorre las columnas para contar entradas y salidas
     for col in dataframe.iter_cols(0, dataframe.max_column):
          
          
          if "X" in col[0].value:
               entradas+=1;
          else:
               if "S" in col[0].value:
                    salidas+=1;
               else:
                    valoresNulos+=1;
     # variable de la cantidad de patrones
     patrones=0;
     # recorre las filas para contar los patrones
     for row in range(1, dataframe.max_row):
          patrones+=1;

     return entradas,salidas,patrones
     
                    
          

     #data.append(_row)

     #print(tabulate(data))

     #print(tabulate(data,headers=encabezado))
     
     
     
readData(ruta)
     